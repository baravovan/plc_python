from flask import request, Flask, send_file, Response
from apscheduler.schedulers.background import BackgroundScheduler
from flask_cors import CORS, cross_origin
from datetime import datetime
from pylogix import PLC

import json
import pytz
import sqlite3
import os.path
import pandas
import openpyxl

# SETTINGS _________________________________________________________________________________________________________________________

dbName = "plcData.db"
timezone='America/Chicago'
 
readPlcTagsIntervalSeconds = 1  #   seconds interval for reading tag values from PLC
cleanDBIntervalDays = 7         #   days interval to delete all tag values
discoverPLCIntervalSeconds = 3  #   seconds interval for discovering the PLC
plcID = 0
plcIpadress = '192.168.4.51'    #   PLC IP adress

# MAIN PROGRAM PART1 ____________________________________________________________________________________________________________________

app = Flask(__name__)

cors = CORS(app)
app.config['CORS_HEADERS'] = 'Content-Type'
app.config["DEBUG"] = True

# APIs _______________________________________________________________________________________________________________________________

@app.route('/', methods=['GET'])
@cross_origin()
def home():
    plc = getPlcFromDB()
    if plc:
        response = json.dumps(plc)
    else:
        response = Response("PLC not found in DB.", status = 400)
    return response

@app.route('/v1/ping', methods=['GET'])
@cross_origin()
def pingAPI():
    plc = ping()
    if plc.DeviceID:
        response = Response("OK", status = 200)
    else:
        response = Response("PLC not found. Check connection.", status = 400)
    return response

# Retreives the controllers tag list
@app.route('/v1/plcTagslist', methods=['GET'])
@cross_origin()
def getPlcTagsList():
    tags = getTagsFromPlc()
    if tags:
        response = json.dumps([ob.__dict__ for ob in tags.Value])
    else:
        response = Response("Tags from PLC not found. Check connection to the PLC.", status = 400)
    return response

# Retrieves only a list of the program names. 
@app.route('/v1/plcProgramsList', methods=['GET'])
@cross_origin()
def getPlcProgramsList():
    programs = getProgramsFromPlc()
    if programs:
        response = json.dumps(programs.Value)
    else:
        response = Response("Programs from PLC not found. Check connection to the PLC.", status = 400)
    return response

# Retrieves a list of a particular program 
@app.route('/v1/plcProgramTagList/<string:name>', methods=['GET'])
@cross_origin()
def getPlcProgramTagList(name):
    tags = getProgramTagList(name)
    if tags:
        response = json.dumps([ob.__dict__ for ob in tags.Value])
    else:
        response = Response("Tags from PLC for program " + name + "not found. Check connection to the PLC.", 
                            status = 400)
    return response

# get latest tag values from DB
@app.route('/v1/getTags', methods=['GET'])
@cross_origin()
def getTags():
    tags = getLatestTagsValuesFromDB()
    if tags:
        response = json.dumps(dict(tags))
    else:
        response = Response("Tag values from DB not found. Check connection to the DB.", status = 400)
    return response

# add Tag
@app.route('/v1/addtag', methods=['POST'])
@cross_origin()
def addTag():
    payload = request.get_json()
    tagName = payload["tag_name"]
    addTagToDB(tagName)
    return "Success", 201

# delete Tag
@app.route('/v1/deleteTag/<string:name>', methods=['DELETE'])
def deleteTag(name):
    print('DELETE tag=' + name)
    deleteTagFromDB(name)
    return "Success", 200

# save to excel OSI Special
@app.route('/v1/saveToExcel', methods=['GET'])
def saveToExcel():
    con = sqlite3.connect(db_path)
    cursor = con.cursor()
    datetime_current = datetime.now(timez)
    filepath = "plc_" + plcID + "_" + datetime_current.strftime("%Y-%m-%d_%H-%M") + ".xlsx"
    #create file
    df = pandas.DataFrame([[plcID, plcIpadress]], 
                          index=['-'], columns=['PLC ID', 'IP Adress'])
    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pandas.ExcelWriter(filepath, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Welcome')

    # Get the xlsxwriter workbook and worksheet objects.
    workbook  = writer.book
    worksheet = writer.sheets['Welcome']

    # Insert an image.
    worksheet.insert_image(3,3, 'logo.png')
    writer._save()

    for i in range(1, 4):
        cursor.execute("SELECT sp, sp_offset, loaded_lb, loaded_sec, load_time FROM v_buckets_OSI_special WHERE bucket_name='Bucket" + str(i) + "'")
        values = cursor.fetchall()
        wb = openpyxl.load_workbook(filepath)
        worksheet = wb.create_sheet("Bucket" + str(i)) # Get sheet
        worksheet.append(['SP lb', 'SP Offset', 'Loaded lb', 'Loaded sec', 'Loaded time'])
        for val in values:
            worksheet.append([val[0], val[1], val[2], val[3], val[4]])
        wb.save(filepath) 
    con.close()
    return send_file(filepath, as_attachment=True)

"""
# save to excel
@app.route('/v1/saveToExcel', methods=['GET'])
def saveToExcel():
    con = sqlite3.connect(db_path)
    cursor = con.cursor()
    cursor.execute("SELECT id, tag_name  FROM tags WHERE plc_id =" + str(plcID))
    tags = cursor.fetchall()
    datetime_current = datetime.now(timez)
    filepath = "plc_" + plcID + "_" + datetime_current.strftime("%Y-%m-%d_%H-%M") + ".xlsx"
    #create file
    df = pandas.DataFrame([[plcID, plcIpadress]], 
                          index=['-'], columns=['PLC ID', 'IP Adress'])
    df.to_excel(filepath, index=False)
    for tag in tags:
        cursor.execute("SELECT value, Timestamp FROM tag_value WHERE tag_id=" + str(tag[0]))
        values = cursor.fetchall()
        wb = openpyxl.load_workbook(filepath)
        work_sheet = wb.create_sheet(removeSpecialChars(tag[1])) # Get sheet
        work_sheet.append(['Timestamp', 'Value'])
        for val in values:
            work_sheet.append([val[1], val[0]])
        wb.save(filepath) 
    con.close()
    return send_file(filepath, as_attachment=True)
"""
def removeSpecialChars(s):
    return s.translate({ord(c): " " for c in "!@#$%^&*()[]{};:,./<>?|~-=_+"})

# errors handling
@app.errorhandler(404)
def page_not_found(e):
    response = json.dumps({"Error":"The resource could not be found."})
    return response

# PLC __________________________________________________________________________________________________________________________________

def ping():
    with PLC(plcIpadress) as comm:
        return comm.GetDeviceProperties().Value

def discover():
    with PLC(plcIpadress) as comm:
        plc = comm.GetDeviceProperties().Value
        savePLC(plc)
        plcID = plc.DeviceID
        discoverScheduler.remove_job('discover_job_id')
        print("PLC discovered. PLC ID: " + str(plcID) + " Ip Address:" + plcIpadress)
    """
    plc_connection = PLC('192.168.4.51')
    plc = plc_connection.Discover()
    plc_connection.Close()
    if len(plc.Value) > 0:
        savePLC(plc.Value[0])
        plcIpadress = plc.Value[0].IPAddress
        plcID = plc.Value[0].DeviceID
        discoverScheduler.remove_job('discover_job_id')
        print("PLC discovered. PLC ID: " + str(plcID) + " Ip Address:" + plcIpadress)
    else:
        print("No reachable PLC discovered. Check network settings.")
    """

def getTagsFromPlc():
    plc_connection = PLC()
    plc_connection.IPAddress = plcIpadress
    tags = plc_connection.GetTagList()
    plc_connection.Close()
    return tags

def readTagsValuesFromPlc(tagsToRead):
    with PLC(plcIpadress) as comm:
        tags = comm.Read(tagsToRead)
        return tags

def getProgramsFromPlc():
    plc_connection = PLC()
    plc_connection.IPAddress = plcIpadress
    programs = plc_connection.GetProgramsList()
    plc_connection.Close()
    return programs

def getProgramTagList(program):
    plc_connection = PLC()
    plc_connection.IPAddress = plcIpadress
    tags = plc_connection.GetProgramTagList(program)
    plc_connection.Close()
    return tags

# DATABASE _____________________________________________________________________________________________________________________________

def savePLC(device):
    con = sqlite3.connect(db_path)
    cursor = con.cursor()
    cursor.execute('SELECT * FROM plc WHERE device_id=' + str(device.DeviceID))
    entry = cursor.fetchone()
    if entry is None:
        con.execute('INSERT INTO plc (device_id, product_name, ipaddress, revision, vendor_id, vendor, device_type, product_code, status, serial, state) VALUES (?,?,?,?,?,?,?,?,?,?,?)', 
                    (str(device.DeviceID), str(device.ProductName), str(device.IPAddress),
                        str(device.Revision), str(device.VendorID), str(device.Vendor),
                        str(device.DeviceType), str(device.ProductCode), str(device.Status),
                        str(device.SerialNumber), str(device.State)
                        ))
        con.commit()
    con.close()

def getPlcFromDB():
    con = sqlite3.connect(db_path)
    cursor = con.cursor()
    cursor.execute('SELECT * FROM plc')
    plc = cursor.fetchone()
    con.close()
    return plc

def getTagsFromDB():
    con = sqlite3.connect(db_path)
    cursor = con.cursor()
    cursor.execute('SELECT * FROM tags WHERE plc_id=' + str(plcID))
    tags = cursor.fetchall()
    con.close()
    return tags

def cleanDB():
    con = sqlite3.connect(db_path)
    con.execute("DELETE FROM tag_value")
    con.commit()
    con.close()

def readAndUpdateTagsValues():
    if plcID != 0 :
        con = sqlite3.connect(db_path)
        cursor = con.cursor()
        cursor.execute('SELECT * FROM tags WHERE plc_id=' + str(plcID))
        tagsToRead = cursor.fetchall()
        con.commit()
        tags = readTagsValuesFromPlc(buildTagNamesList(tagsToRead))
        for tag in tags:
            cursor.execute("SELECT id FROM tags WHERE plc_id=" + str(plcID) + " AND tag_name='" + tag.TagName + "'")
            id = cursor.fetchone()
            con.execute('INSERT INTO tag_value (value, tag_id) VALUES (?,?)', (tag.Value, id[0]))
            con.commit()
        con.close()

def readAndUpdateTagsValuesOsiSpecial():
    bucketStatuses = readTagsValuesFromPlc(["Bucket1_TrackSeq.Status", "Bucket2_TrackSeq.Status", "Bucket3_TrackSeq.Status"])
    if bucketStatuses[0].Value == 100:
        bucketString = 'Bucket1'
    elif bucketStatuses[1].Value == 100:
        bucketString = 'Bucket2'
    elif bucketStatuses[2].Value == 100:
        bucketString = 'Bucket3'
    else:
        bucketString = ''
    
    if bucketString != '':
        con = sqlite3.connect(db_path)
        cursor = con.cursor()
        cursor.execute("SELECT * FROM tags WHERE plc_id=" + str(plcID) + " AND description='" + bucketString + "'")
        tagsToRead = cursor.fetchall()
        con.commit()
        tags = readTagsValuesFromPlc(buildTagNamesList(tagsToRead))
        for tag in tags:
            cursor.execute("SELECT id FROM tags WHERE plc_id=" + str(plcID) + " AND tag_name='" + tag.TagName + "'")
            id = cursor.fetchone()
            con.execute('INSERT INTO tag_value (value, tag_id) VALUES (?,?)', (tag.Value, id[0]))
            con.commit()
        con.close()

def buildTagNamesList(tagsToRead):
    list = [];
    for tag in tagsToRead:
        list.append(tag[1])
    return list

def getLatestTagsValuesFromDB():
    con = sqlite3.connect(db_path)
    cursor = con.cursor()
    cursor.execute('SELECT tag_name, value FROM tag_value as v LEFT JOIN tags as t ON v.tag_id = t.id WHERE plc_id=' + str(plcID) + ' ORDER BY "Timestamp" DESC LIMIT (SELECT count() FROM tags WHERE plc_id=' + str(plcID) + ')');
    tags = cursor.fetchall()
    con.close()
    return tags

def addTagToDB(tagName):
    con = sqlite3.connect(db_path)
    plc = getPlcFromDB()
    con.execute('INSERT INTO tags (tag_name, plc_id) VALUES (?,?)', (tagName, str( plc[1])))
    con.commit()
    con.close()

def deleteTagFromDB(name):
    con = sqlite3.connect(db_path)
    cursor = con.cursor()
    cursor.execute("SELECT id FROM tags WHERE tag_name='" + name + "'")
    id = cursor.fetchone()
    con.execute("DELETE FROM tags WHERE tag_name='" + name + "'")
    con.execute("DELETE FROM tag_value WHERE tag_id=" + str(id[0]))
    con.commit()
    con.close()

# MAIN PROGRAM PART2 ____________________________________________________________________________________________________________________

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
db_path = os.path.join(BASE_DIR, dbName)
timez= pytz.timezone(timezone)
plc = getPlcFromDB()
if plc:
    plcID = plc[1]
    plcIpadress = plc[3]
else:
    discoverScheduler = BackgroundScheduler(timezone=timezone)
    discoverScheduler.add_job(discover, 'interval', seconds=discoverPLCIntervalSeconds, id='discover_job_id')
    discoverScheduler.start()
   

cleanScheduler = BackgroundScheduler(timezone=timezone)
cleanScheduler.add_job(cleanDB, 'interval', days=cleanDBIntervalDays)

readScheduler = BackgroundScheduler(timezone=timezone, job_defaults={'max_instances': 2})
readScheduler.add_job(readAndUpdateTagsValuesOsiSpecial, 'interval', seconds=readPlcTagsIntervalSeconds)

cleanScheduler.start()
readScheduler.start()

try:
    #if __name__ == '__main__':
     #   app.run(host='0.0.0.0', port=5000, use_reloader=False)
    app.run(use_reloader=False)
except (KeyboardInterrupt, SystemExit):
    cleanScheduler.shutdown()
    readScheduler.shutdown()
    #discoverScheduler.shutdown()